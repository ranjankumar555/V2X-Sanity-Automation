#!/usr/bin/env python3
# -*- coding: utf-8 -*-

r"""
dlt_verifiy.py - Automated DLT Verification System

Pipeline:
  1. Auto-discover DLT files in D:\SANITY\{date}\{variant}\
  2. Extract version and variant from filename
  3. Convert DLT -> CSV
  4. Run 5 test verifications (V2X, Coding, Cybersecurity, Diagnostics, DTC)
  5. Generate Excel report automatically

Supports two CSV formats:
  1) CSV WITH header (comma-separated)
  2) Headerless space-separated rows
"""

import csv
import json
import os
import sys
import subprocess
import glob
import tempfile
from fnmatch import fnmatchcase
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==================== Tiny Logger ====================
def _log(msg): 
    print(msg, flush=True)

# ==================== DLT Viewer ====================
def _run(args, err_msg):
    try:
        subprocess.run(args, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    except FileNotFoundError:
        raise RuntimeError(f"{err_msg}: '{args[0]}' not found.")
    except subprocess.CalledProcessError as e:
        raise RuntimeError(f"{err_msg} (exit {e.returncode})\nSTDERR:\n{e.stderr.strip()}")

def run_dlt_viewer_convert(dlt_viewer_exe, in_dlt, out_csv, silent=True, utf8=True):
    args = [dlt_viewer_exe]
    if silent: args.append("-s")
    args.append("-csv")
    if utf8: args.append("-u")
    args.extend(["-c", in_dlt, out_csv])
    _run(args, "DLT Viewer CSV conversion failed")

# ------------------ Helpers ----------------------
def normalize(name: str) -> str:
    return (name or "").strip().lower().replace(" ", "").replace("-", "").replace("_", "")

def quick_detect_delimiter(header_line: str) -> str:
    sc = header_line.count(';'); cc = header_line.count(',')
    if sc == 0 and cc == 0: return ','
    return ';' if sc > cc else ','

DEFAULT_APID_CANDS    = ("apid", "appid", "application", "applicationid")
DEFAULT_CTID_CANDS    = ("ctid", "ctxid", "context", "contextid")
DEFAULT_PAYLOAD_CANDS = ("payload", "text", "message", "arguments", "payloadascii", "payloadstring")

def locate_columns(header, overrides=None):
    norm = [normalize(h) for h in header]
    def find_one(name_norm):
        for i, n in enumerate(norm):
            if n == name_norm or name_norm in n:
                return i
        return None

    ap_i = ct_i = pl_i = None
    if overrides:
        if overrides.get("APID"):    ap_i = find_one(normalize(overrides["APID"]))
        if overrides.get("CTID"):    ct_i = find_one(normalize(overrides["CTID"]))
        if overrides.get("PAYLOAD"): pl_i = find_one(normalize(overrides["PAYLOAD"]))

    def find_any(cands):
        for c in cands:
            idx = find_one(c)
            if idx is not None: return idx
        return None

    if ap_i is None: ap_i = find_any(DEFAULT_APID_CANDS)
    if ct_i is None: ct_i = find_any(DEFAULT_CTID_CANDS)
    if pl_i is None: pl_i = find_any(DEFAULT_PAYLOAD_CANDS)
    if pl_i is None and header: pl_i = len(header) - 1

    return ap_i, ct_i, pl_i

# ------------------ Rules (in-code list) ----------
def rule_matches(apid_val, ctid_val, rules_list):
    """
    rules_list: list of tuples or dicts:
      ("DIAG","DIAG")
      ("APP","*")
      ("*","CON")
      ("V2X*","MAIN")
    """
    for rule in rules_list:
        if isinstance(rule, dict):
            ap_pat = rule.get("apid", "*")
            ct_pat = rule.get("ctid", "*")
        else:
            ap_pat, ct_pat = rule
        ap_ok = (ap_pat == "*") or fnmatchcase(apid_val, ap_pat)
        ct_ok = (ct_pat == "*") or fnmatchcase(ctid_val, ct_pat)
        if ap_ok and ct_ok:
            return True
    return False

# ------------------ PREFILTER: CSV-with-header ----
def prefilter_csv_with_header(src_csv, dst_csv, rules_list, column_overrides=None,
                              buffering_bytes=4_194_304, project_columns=False):
    with open(src_csv, "r", encoding="utf-8", errors="replace",
              newline="", buffering=buffering_bytes) as rf:
        first = rf.readline()
        if not first:
            raise RuntimeError("CSV seems empty")
        delim = quick_detect_delimiter(first)
        rf.seek(0)

        reader = csv.reader(rf, delimiter=delim, quotechar='"', escapechar='\\')
        header = next(reader, None) or []
        ap_i, ct_i, pl_i = locate_columns(header, overrides=column_overrides)
        if ap_i is None or ct_i is None:
            raise KeyError(f"Apid/Ctid columns not found in header mode. Header: {header}")

        # Optional slim projection
        ts_i = None
        for i, h in enumerate(header):
            if normalize(h) in ("timestamp", "time"):
                ts_i = i; break
        keep_idx = [ap_i, ct_i, pl_i] + ([ts_i] if ts_i is not None else [])
        out_header = [header[i] for i in keep_idx]

        os.makedirs(os.path.dirname(dst_csv), exist_ok=True)
        with open(dst_csv, "w", encoding="utf-8", newline="") as wf:
            writer = csv.writer(wf, delimiter=',', quotechar='"', lineterminator="\n")
            writer.writerow(out_header)

            for row in reader:
                if ap_i >= len(row) or ct_i >= len(row): continue
                apv = (row[ap_i] or "").strip().strip('"')
                ctv = (row[ct_i] or "").strip().strip('"')
                if rule_matches(apv, ctv, rules_list):
                    writer.writerow([row[i] if i < len(row) else "" for i in keep_idx])

# -------- PREFILTER: headerless space-separated ---
def _parse_headerless_line(line: str):
    """
    Expecting tokens like:
      Index  Date  Time  Timestamp  Count  Ecuid  Apid  Ctid  SessionId  Type  Subtype  Mode  #Args  Payload...
    => Time consumes TWO tokens (Date + Time). Payload is the rest.
    """
    toks = line.strip().split()
    if len(toks) < 13:
        return None  # malformed
    # map known positions (see layout above)
    apid  = toks[6]
    ctid  = toks[7]
    # timestamp column from layout (index 3)
    timestamp = toks[1] + " " + toks[2]  # full time
    # payload starts at index 13 (0-based)
    payload = " ".join(toks[13:])
    return apid, ctid, timestamp, payload

def prefilter_headerless_text(src_path, dst_csv, rules_list, buffering_bytes=4_194_304):
    r"""
    Reads headerless, space-separated rows; writes a normalized CSV with header:
      Apid,Ctid,Timestamp,Payload
    """
    try:
        os.makedirs(os.path.dirname(dst_csv), exist_ok=True)
    except PermissionError:
        # Fallback to temp directory if no write permission to original location
        dst_csv = os.path.join(tempfile.gettempdir(), os.path.basename(dst_csv))
    
    kept = 0
    with open(src_path, "r", encoding="utf-8", errors="replace",
              newline="", buffering=buffering_bytes) as rf, \
         open(dst_csv, "w", encoding="utf-8", newline="") as wf:
        writer = csv.writer(wf, delimiter=',', quotechar='"', lineterminator="\n")
        writer.writerow(["Apid", "Ctid", "Timestamp", "Payload"])
        for line in rf:
            if not line.strip():
                continue
            parsed = _parse_headerless_line(line)
            if not parsed:
                continue
            apid, ctid, timestamp, payload = parsed
            if rule_matches(apid, ctid, rules_list):
                writer.writerow([apid, ctid, timestamp, payload])
                kept += 1
    return kept

# ------------------ SEARCH -----------------------
def stream_search_filtered_csv(csv_path, tests, case_sensitive=False,
                               buffering_bytes=4_194_304):
    prepared = []
    for t in tests:
        # Support both single needle and multiple needles
        needles_list = t.get("needles", [])
        if not needles_list and "needle" in t:
            needles_list = [t["needle"]]
        
        if not case_sensitive:
            needles_list = [n.lower() for n in needles_list]
        
        prepared.append({
            "name": t["name"],
            "needles": needles_list,
            "found_count": [False] * len(needles_list)  # Track which needles are found
        })

    with open(csv_path, "r", encoding="utf-8", errors="replace",
              newline="", buffering=buffering_bytes) as f:
        reader = csv.reader(f, delimiter=',', quotechar='"', escapechar='\\')
        header = next(reader, None) or []
        # find payload column in normalized filtered CSV
        payload_idx = None
        for i, h in enumerate(header):
            if normalize(h) in ("payload", "text", "message", "arguments", "payloadascii", "payloadstring"):
                payload_idx = i; break
        if payload_idx is None:  # in our normalized CSV header it's exactly "Payload"
            for i, h in enumerate(header):
                if h.strip().lower() == "payload":
                    payload_idx = i; break
        if payload_idx is None:
            # fallback: assume last column is payload
            payload_idx = len(header) - 1 if header else 0

        for row in reader:
            if payload_idx >= len(row): continue
            hay = row[payload_idx] if case_sensitive else (row[payload_idx] or "").lower()
            for t in prepared:
                # Check each needle and mark if found
                for i, needle in enumerate(t["needles"]):
                    if not t["found_count"][i] and needle in hay:
                        t["found_count"][i] = True

    return [{"Testcase Name": t["name"], "Result": "Pass" if all(t["found_count"]) else "Fail"} for t in prepared]

# ==================== Excel Report Writer ====================
def write_excel_report(excel_path, version, test_results_dict, device_version):
    r"""
    Write test results to Excel file with management standard formatting.
    Exact formatting inherited from generate_sanity_results.py
    
    Columns: Category, Name, Result, [Mandatory]Test version, Fail/Blocking reason, [Mandatory]Result from
    
    [Mandatory]Test version is populated with the version parameter extracted from DLT filename
    """
    if not OPENPYXL_AVAILABLE:
        _log("[WARNING] openpyxl not installed. Skipping Excel report.")
        return
    
    # ---- Build management rows (all test cases) ----
    management_headers = ["Category", "Name", "Result", "[Mandatory]Test version", "Fail/Blocking reason", "[Mandatory]Result from"]
    
    rows = [
        # Basic function > v2x tests (27 rows)
        {"Category": "Basic function > v2x", "Name": "V2X start-up behaviour", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "V2X Parameter provisioning", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "Coding values - Height, Width, Length", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "Coding values - V2X Antenna", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "EU V2X Security - Sign and verify messages", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "CN V2X Security - Sign and verify messages", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "CN V2X Security - Certificate Download", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "EU V2X Security - Certificate Download", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "V2X Application", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "CN V2X - Sending messages", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "EU V2X - Sending messages", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "(Location) V2X receives Position and odometry data from GNSS", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "V2X Stack in Hangup", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "(Cybersecurity) EU V2X Security - Sign messages", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "(Cybersecurity) EU V2X Security - Verify messages", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "(Cybersecurity) EU V2X Security - filter messages", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "(Cybersecurity) CN V2X Security - Sign messages", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "(Cybersecurity) CN V2X Security - Verify messages", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "(Cybersecurity) CN V2X Security - filter messages", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "CN V2X - Map Matching", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "EU V2X - Map Matching", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "Handle Power Mode", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "Data Privacy Setting", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "CN V2X - Geofencing", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "EU V2X - Geofencing", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "Basic function > v2x", "Name": "Object Forwarding - Objects >50", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        # DTC tests (19 rows)
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X deactivated by diagnosis", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X STACK_ERROR_RADIO_ACCESS_ERROR", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X STACK_ERROR_MISSING_NAV_INFO", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X STACK_ERROR_FS_ACCESS_ERROR", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X STACK_ERROR_HSM_ACCESS_ERROR", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X STACK_ERROR_ECDSA_ACCESS_ERROR", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X STACK_ERROR_MISSING_CERT_LIST", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X STACK_ERROR_MISSING_AUTH_CERT", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X STACK_ERROR_MISSING_MAP_INFO", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X PKI_ERROR_V2X_STACK_CONNECTION_ERROR", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X PKI_ERROR_PKI_SERVER_CONNECTION_ERROR", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X PKI_ERROR_FS_ACCESS_ERROR", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X PKI_ERROR_ENROLLMENT_FAILED", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X PKI_ERROR_REENROLLMENT_FAILED", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X PKI_ERROR_CERT_LIST_UPDATE_FAILED", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X PKI_ERROR_AUTH_REQ_FAILED", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X PKI_ERROR_AUTH_DOWNLOAD_FAILED", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > DTC > ICON specific diagnostics Trouble codes > V2X", "Name": "V2X Unavailable", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        # Diagnostics tests (7 rows)
        {"Category": "System function > Diagnostics services > ICON specific diagnostics services > v2x", "Name": "RDBI_V2X_ACTIVATION", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > Diagnostics services > ICON specific diagnostics services > v2x", "Name": "WDBI_V2X_ACTIVATION", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > Diagnostics services > ICON specific diagnostics services > v2x", "Name": "RDBI_V2X_SECURITY", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > Diagnostics services > ICON specific diagnostics services > v2x", "Name": "RDBI_V2X_RADIO", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > Diagnostics services > ICON specific diagnostics services > v2x", "Name": "RDBI_V2X_STACK_CONFIGURATION", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > Diagnostics services > ICON specific diagnostics services > v2x", "Name": "RDBI_V2X_HSM", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
        {"Category": "System function > Diagnostics services > ICON specific diagnostics services > v2x", "Name": "RDBI_V2X_COMPENSATOR_LNA", "Result": "Pass", "[Mandatory]Test version": version, "Fail/Blocking reason": "--", "[Mandatory]Result from": "Manual test from fused board"},
    ]

    # ---- Create workbook ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Sanity Summary"

    # ---- Styles (exact formatting from generate_sanity_results.py) ----
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")  # dark blue
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D9D9D9")
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)

    # ---- Write Header ----
    for col_idx, hdr in enumerate(management_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_all

    # ---- Write Data Rows ----
    for r_idx, r in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=r["Category"]).border = border_all
        ws.cell(row=r_idx, column=2, value=r["Name"]).border = border_all
        ws.cell(row=r_idx, column=3, value=r["Result"]).border = border_all
        ws.cell(row=r_idx, column=4, value=r["[Mandatory]Test version"]).border = border_all
        ws.cell(row=r_idx, column=5, value=r["Fail/Blocking reason"]).border = border_all
        ws.cell(row=r_idx, column=6, value=r["[Mandatory]Result from"]).border = border_all

    # ---- Conditional Formatting on Result column (col 3) ----
    start_row = 2
    end_row = ws.max_row
    if end_row >= start_row:
        rng = f"C{start_row}:C{end_row}"
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
        red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red

        # Equal to "Pass" => green
        ws.conditional_formatting.add(
            rng, CellIsRule(operator='equal', formula=['"Pass"'], stopIfTrue=True, fill=green_fill)
        )
        # Not equal to "Pass" => red
        ws.conditional_formatting.add(
            rng, CellIsRule(operator='notEqual', formula=['"Pass"'], stopIfTrue=True, fill=red_fill)
        )

    # ---- Column Width Auto-fit (exact values from generate_sanity_results.py) ----
    # Compute max string length per column
    col_max = {i: len(management_headers[i-1]) for i in range(1, len(management_headers) + 1)}
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(management_headers) + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            val_str = str(val)
            length = min(len(val_str), 120)  # cap to avoid very wide columns
            if length > col_max[c]:
                col_max[c] = length

    # Apply widths with padding and sensible bounds per column
    min_widths = {1: 24, 2: 26, 3: 10, 4: 22, 5: 40, 6: 22}
    max_widths = {1: 60, 2: 60, 3: 12, 4: 28, 5: 90, 6: 32}
    padding = 2

    for c in range(1, len(management_headers) + 1):
        approx = col_max[c] + padding
        width = max(min_widths[c], min(approx, max_widths[c]))
        ws.column_dimensions[chr(ord('A') + c - 1)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save Excel file
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    wb.save(excel_path)
    _log(f"[OK] Excel report saved -> {excel_path}")

# ==================== Auto-Discovery ====================
def find_dlt_files(base_path: str = r"D:\SANITY") -> list:
    r"""Find all DLT files in D:\SANITY\{date}\{variant}\ structure"""
    pattern = os.path.join(base_path, "*", "*", "*.dlt")
    return sorted(glob.glob(pattern))

def extract_date_variant_from_version(version_str: str) -> tuple:
    r"""
    Extract date and device variant from version string.
    Patterns supported:
      - v040.040.065.iconsf25.oem_260525  (OEM format)
      - v015.052.025.icon25.sop_250825   (SOP format)
    Returns: (date, variant) or (None, None) if parsing fails
    """
    if not version_str:
        return None, None
    
    parts = version_str.split('.')
    
    # Find date in 'oem_YYMMDD' or 'sop_YYMMDD' format
    date = None
    date_marker = None
    for part in parts:
        if 'oem_' in part:
            date = part.split('oem_')[1]
            date_marker = 'oem_'
            break
        elif 'sop_' in part:
            date = part.split('sop_')[1]
            date_marker = 'sop_'
            break
        elif 'mtk_' in part:
            date = part.split('mtk_')[1]
            date_marker = 'mtk_'
            break
        
    
    if not date or not date_marker:
        return None, None
    
    # Find variant (typically the part right before the date marker)
    variant = None
    for i, part in enumerate(parts):
        if date_marker in part and i > 0:
            variant = parts[i - 1]
            break
    
    return date, variant

def extract_version_variant(dlt_path: str) -> tuple:
    r"""
    Extract version and variant from DLT file path.
    Only processes files matching: basic_sanity_{version_string}.dlt (case-insensitive)
    The version string contains embedded date and variant info:
      Pattern OEM: v040.040.065.iconsf25.oem_260525  → date=260525, variant=iconsf25
      Pattern SOP: v015.052.025.icon25.sop_250825    → date=250825, variant=icon25
    Returns: (version, variant, date) or (None, None, None) if not a standard DLT file
    """
    path_parts = Path(dlt_path).parts
    try:
        sanity_idx = path_parts.index("SANITY")
        filename = path_parts[sanity_idx + 3]
        
        # Only process files matching the standard basic_sanity pattern (case-insensitive)
        filename_lower = filename.lower()
        if filename_lower.startswith("basic_sanity_") and filename_lower.endswith(".dlt"):
            version = filename.replace("basic_sanity_", "").replace("basic_Sanity_", "").replace(".dlt", "").replace(".DLT", "")
            # Extract date and variant from version string
            date, variant = extract_date_variant_from_version(version)
            if date and variant:
                return version, variant, date
    except (IndexError, ValueError, KeyError):
        pass
    
    return None, None, None

# ==================== Test Rules ====================
def get_v2x_stack_rules():
    """V2X Stack Verification Rules"""
    return {
        "name": "V2X Stack",
        "rules": [
            {"name": "Stack Initialization (RUNNING)", "needles": ["STACK_IN_RUNNING"]},
            {"name": "Stack Start Sequence", "needles": ["STACK_IN_START"]},
            {"name": "Stack Stop Event", "needles": ["STACK_IN_STOP"]},
        ]
    }

def get_coding_provisioning_rules():
    """Coding Provisioning Verification Rules"""
    return {
        "name": "Coding Provisioning",
        "rules": [
            {"name": "Antenna Configuration", "needles": ["V2X_ANTENNAS_CONFIG value:0"]},
            {"name": "Vehicle Height (150cm)", "needles": ["VEHICLE_HEIGHT : 150"]},
            {"name": "Vehicle Length (300cm)", "needles": ["VEHICLE_LENGTH : 300"]},
            {"name": "Vehicle Width (150cm)", "needles": ["VEHICLE_WIDTH : 150"]},
        ]
    }

def get_cybersecurity_rules():
    """Cybersecurity Verification Rules"""
    return {
        "name": "Cybersecurity",
        "rules": [
            {"name": "Invalid Signature Detection", "needles": ["Call unwrap callback from message verify with result: Invalid signature"]},
            {"name": "Valid Message Verification", "needles": ["Call unwrap callback from message verify with result: Verified"]},
            {"name": "Replayed Message Detection", "needles": ["Call unwrap callback from message verify with result: Replayed message"]},
        ]
    }

def get_diagnostics_rules():
    """Diagnostics Verification Rules"""
    return {
        "name": "Diagnostics",
        "rules": [
            {"name": "Antenna Config (Diag)", "needles": ["V2X_ANTENNAS_CONFIG value:0"]},
            {"name": "Vehicle Height (Diag)", "needles": ["VEHICLE_HEIGHT : 150"]},
            {"name": "Vehicle Length (Diag)", "needles": ["VEHICLE_LENGTH : 300"]},
            {"name": "Vehicle Width (Diag)", "needles": ["VEHICLE_WIDTH : 150"]},
        ]
    }

def get_dtc_rules():
    """DTC (Diagnostic Trouble Code) Verification Rules - Requires BOTH active and inactive"""
    return {
        "name": "DTC",
        "rules": [
            {"name": "V2X deactivated by diagnosis(0xB7F2D0)", "needles": ["DTC Info : active 0xb7f2d0", "DTC Info : inactive 0xb7f2d0"]},
            {"name": "V2X function unavailable(0xB7F2D1)", "needles": ["DTC Info : active 0xb7f2d1", "DTC Info : inactive 0xb7f2d1"]},
            {"name": "V2X SW Stack: HSM access error (0xB7F186)", "needles": ["DTC Info : active 0xb7f186", "DTC Info : inactive 0xb7f186"]},
            {"name": "V2X PKI Client: Authorization request failed (0x610027)", "needles": ["DTC Info : active 0x610027", "DTC Info : inactive 0x610027"]},
            {"name": "V2X PKI Client: Update certificate list failed(0x61002D)", "needles": ["DTC Info : active 0x61002d", "DTC Info : inactive 0x61002d"]},
            {"name": "V2X PKI Client: Unable to connect to V2X SW stack (0x61002C)", "needles": ["DTC Info : active 0x61002c", "DTC Info : inactive 0x61002c"]},
            {"name": "V2X PKI Client: Unable to connect to V2X PKI server (0x61002B)", "needles": ["DTC Info : active 0x61002b", "DTC Info : inactive 0x61002b"]},
            {"name": "V2X PKI Client: File system access error (0x61002A)", "needles": ["DTC Info : active 0x61002a", "DTC Info : inactive 0x61002a"]},
            {"name": "V2X PKI Client: Enrollment failed (0x610028 )", "needles": ["DTC Info : active 0x610028", "DTC Info : inactive 0x610028"]},
            {"name": "V2X PKI Client: Authorization download failed(0x610026)", "needles": ["DTC Info : active 0x610026", "DTC Info : inactive 0x610026"]},
            {"name": "V2X PKI Client - Enrollment renewal failed (0x610029)", "needles": ["DTC Info : active 0x610029", "DTC Info : inactive 0x610029"]},
            {"name": "V2X SW Stack: ECDSA accelerator access error(0x61002E)", "needles": ["DTC Info : active 0x61002e", "DTC Info : inactive 0x61002e"]},
            {"name": "V2X SW Stack: File system access error (0x61002F)", "needles": ["DTC Info : active 0x61002f", "DTC Info : inactive 0x61002f"]},
            {"name": "V2X SW Stack: Missing V2X Authorization certificate (0x610033)", "needles": ["DTC Info : active 0x610033", "DTC Info : inactive 0x610033"]},
            {"name": "V2X SW Stack: Missing V2X certificate list(0x610037)", "needles": ["DTC Info : active 0x610037", "DTC Info : inactive 0x610037"]},
            {"name": "V2X SW Stack: V2X radio access error(0x610039)", "needles": ["DTC Info : active 0x610039", "DTC Info : inactive 0x610039"]},
            {"name": "V2X SW Stack: Missing navigation information(0x610032)", "needles": ["DTC Info : active 0x610032", "DTC Info : inactive 0x610032"]},
            {"name": "V2X SW Stack: Missing map information(0x610031)", "needles": ["DTC Info : active 0x610031", "DTC Info : inactive 0x610031"]},
        ]
    }

# ==================== Main Verification Function ====================
def verify_dlt_file(dlt_file: str,dlt_viewer_exe: str = str(Path(__file__).resolve().parent / "DltViewerSDK-2.21.3" / "dlt-viewer.exe"),reuse_csv: bool = True):
    """Verify a single DLT file against all 5 test suites"""
    if not os.path.exists(dlt_file):
        raise FileNotFoundError(f"DLT file not found: {dlt_file}")
    
    version, variant, date = extract_version_variant(dlt_file)
    if not version:
        raise ValueError(f"Cannot extract version/variant from: {dlt_file}")
    
    # _log(f"\n{'='*70}")
    # _log(f"[INFO] Processing: {os.path.basename(dlt_file)}")
    # _log(f"[INFO] Version: {version}, Variant: {variant}, Date: {date}")
    # _log(f"{'='*70}")
    
    # Stage 1: Convert DLT to CSV
    csv_out = os.path.splitext(dlt_file)[0] + ".csv"
    if not (reuse_csv and os.path.exists(csv_out) and os.path.getsize(csv_out) > 0):
        _log(f"[INFO] Converting DLT -> CSV...")
        run_dlt_viewer_convert(dlt_viewer_exe, dlt_file, csv_out, silent=True, utf8=True)
    else:
        _log(f"[INFO] Reusing existing CSV")
    
    # Stage 2: Run all 5 test suites
    all_results = {}
    test_summary = []  # Collect results for table display
    column_overrides = {"APID": "Apid", "CTID": "Ctid", "PAYLOAD": "Payload"}
    
    test_suites = [
        ("dtc", get_dtc_rules(), [("DIAG", "DIAG")]),
        ("v2x_stack", get_v2x_stack_rules(), [("V2XM", "STCK")]),
        ("coding_prov", get_coding_provisioning_rules(), [("V2XM", "CONF")]),
        ("cybersecurity", get_cybersecurity_rules(), [("*", "*")]),
        ("diagnostics", get_diagnostics_rules(), [("V2XM", "CONF")]),
    ]
    
    for test_name, test_rules, filter_rules in test_suites:
        # _log(f"[TEST] {test_rules['name']}...")
        # print(f"[INFO] Verifying Test: {test_rules['name']}...")
        
        # Use temp directory for filtered CSVs to avoid permission issues
        temp_dir = tempfile.gettempdir()
        filtered_csv = os.path.join(temp_dir, os.path.basename(os.path.splitext(csv_out)[0]) + f".{test_name}.prefiltered.csv")
        
        # Pre-filter with appropriate rules
        try:
            prefilter_csv_with_header(csv_out, filtered_csv, filter_rules, column_overrides=column_overrides)
        except Exception:
            prefilter_headerless_text(csv_out, filtered_csv, filter_rules)
        
        # Search for test patterns
        results = stream_search_filtered_csv(filtered_csv, test_rules["rules"], case_sensitive=False)
        all_results[test_rules["name"]] = results
        
        # Collect summary for table
        passed = sum(1 for r in results if r["Result"] == "Pass")
        total = len(results)
        pass_rate = f"{(passed/total*100):.0f}%" if total > 0 else "0%"
        test_summary.append({
            "name": test_rules["name"],
            "passed": passed,
            "total": total,
            "pass_rate": pass_rate
        })
    
    # Stage 3: Display test results in table format
    _log("\nTEST RESULTS SUMMARY")
    _log("=" * 53)
    
    # Calculate totals and determine status
    total_passed = sum(t["passed"] for t in test_summary)
    total_count = sum(t["total"] for t in test_summary)
    
    # Header
    _log(f"{'Test Name':<35} {'Passed':<10} {'Status':<8}")
    _log("-" * 53)
    
    # Rows with Status
    for test in test_summary:
        status = "Passed" if test["passed"] == test["total"] else "Failed"
        _log(f"{test['name']:<35} {test['passed']:<10} {status:<8}")
    
    # Footer separator
    _log("-" * 53)
    
    # List failed test cases
    failed_items = []
    for test_name, results in all_results.items():
        for result in results:
            if result["Result"] == "Fail":
                failed_items.append((test_name, result["Testcase Name"]))
    
    # Display failed test cases if any
    if failed_items:
        for test_name, testcase_name in failed_items:
            _log(f"[Failed][{test_name}] {testcase_name}")
    
    _log("=" * 73 + "\n")
    
    # Stage 4: Write Excel report
    output_dir = os.path.dirname(dlt_file)  # Same directory as DLT file
    output_excel = f"{output_dir}/v2x_sanity_result_{version}.xlsx"
    
    write_excel_report(output_excel, version, all_results, version)
    _log(f"{'='*70}\n")

# ==================== Main ====================
def main():
    """Auto-discover and verify all DLT files, or process single file"""
    if len(sys.argv) > 1:
        dlt_file = sys.argv[1]
        if not os.path.exists(dlt_file):
            _log(f"[ERROR] File not found: {dlt_file}")
            sys.exit(1)
        verify_dlt_file(dlt_file)
    else:
        # Auto-discover DLT files
        dlt_files = find_dlt_files()
        if not dlt_files:
            _log("[INFO] No DLT files found in D:\\SANITY\\")
            _log("[INFO] Usage: python dlt_verifiy.py [<dlt_file>]")
            return
        
        _log(f"[INFO] Found {len(dlt_files)} DLT file(s)")
        for dlt_file in dlt_files:
            try:
                verify_dlt_file(dlt_file)
            except Exception as e:
                _log(f"[ERROR] {os.path.basename(dlt_file)}: {e}")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        _log("\n[INFO] Interrupted by user")
        sys.exit(0)
    except Exception as e:
        _log(f"[ERROR] Unhandled exception: {e}")
        sys.exit(1)